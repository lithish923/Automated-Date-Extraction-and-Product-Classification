import os
from object_classification import object_classification
from text_extraction_ocr import ocr
from extract_info import extract_info
from extract_mrp import extract_mrp
from freshness_test import testing_freshness
import openpyxl
from openpyxl import Workbook
from datetime import datetime

packed_file = "packed_product_data.xlsx"
unpacked_file = "unpacked_product_data.xlsx"

# Function to initialize Excel file
def initialize_workbook(file_name, headers):
    if os.path.exists(file_name):
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
        except Exception as e:
            print(f"Error loading {file_name}: {e}")
            os.remove(file_name)
            print(f"Corrupted file {file_name} deleted.")
            wb = Workbook()
            sheet = wb.active
            sheet.append(headers)
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(headers)
    return wb, sheet

# Initialize workbooks for packed and unpacked products
packed_headers = ["Product Number", "Manufacturing Date", "Expiry Date", "Shelf Life", "Quantity", "MRP", "Status"]
unpacked_headers = ["Sl. No", "Name", "Fresh/Rotten"]
packed_wb, packed_sheet = initialize_workbook(packed_file, packed_headers)
unpacked_wb, unpacked_sheet = initialize_workbook(unpacked_file, unpacked_headers)

# Folder containing the images
folder_path = r"test_images"
product_number = packed_sheet.max_row
sl_no = unpacked_sheet.max_row

for file_name in os.listdir(folder_path):
    img_path = os.path.join(folder_path, file_name)

    if file_name.lower().endswith('.jpg'):
        print(f"Processing: {file_name}")

        try:
            classification_result = object_classification(img_path)
            print(classification_result)

            if classification_result == "Images_Packed":
                classification_result = "Packed Item"
                extracted_text = ocr(img_path)
                packed_info = extract_info(extracted_text)
                mrp = extract_mrp(extracted_text)
                if packed_info is not None:
                    packed_info['mrp'] = mrp

                shelf_life = f"{packed_info['shelf_life']['years']}years{packed_info['shelf_life']['months']}months{packed_info['shelf_life']['days']}days" if 'shelf_life' in packed_info else None
                quantity = f"{packed_info['quantities'][0]}{packed_info['quantities'][1]}" if 'quantities' in packed_info else None
                if(packed_info['shelf_life']['years']>0 or packed_info['shelf_life']['months']>0  or  packed_info['shelf_life']['days'] >0):
                    expiry_status="Not Expired"
                else:
                    expiry_status="Expired"


                row = [
                    product_number,
                    packed_info.get('mfg_date', None),
                    packed_info.get('expiry_date', None),
                    shelf_life,
                    quantity,
                    packed_info.get('mrp', 'None'),
                    expiry_status
                ]
                packed_sheet.append(row)
                product_number += 1

                print(packed_info)
            else:
                predicted_class_name, shelf_life = testing_freshness(img_path)

                freshness_status = "Fresh" if predicted_class_name == "Fresh" else "Rotten"

                row = [
                    sl_no,
                    predicted_class_name,
                    freshness_status
                ]
                unpacked_sheet.append(row)
                sl_no += 1

                print({
                    'Name': predicted_class_name,
                    'Fresh/Rotten': freshness_status
                })

            print(f"Deleted: {file_name}")
        except Exception as e:
            print(f"Error processing {file_name}: {e}")

packed_wb.save(packed_file)
unpacked_wb.save(unpacked_file)
print(f"Packed data saved to {packed_file}")
print(f"Unpacked data saved to {unpacked_file}")
